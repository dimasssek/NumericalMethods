import math
from openpyxl import Workbook

def u(t):
    return t + (t / (t + 1))

def f(t, u):
    return (1 / t) * ((2 * t + 1) * u - u * u - t * t)

def average(array_list):
    return sum(array_list) / len(array_list)

abscissa, ordinates, solution = [], [], []
with open("input2.txt") as file:
    t0, b, u0, h = map(float, file.readline().split())
a = t0

t_prev = t0
y_prev = u0

k1 = h * f(t_prev, y_prev)
k2 = h * f(t_prev + h, y_prev + h * k1)
y_curr = y_prev + 0.5 * (k1 + k2)
t_curr = t_prev + h

dot = a
while dot <= b:
    abscissa.append(dot)
    solution.append(u(dot))
    dot += h

ordinates.append(y_prev)
ordinates.append(y_curr)

for i in range(2, len(abscissa)):
    y_curr = y_curr + h * (1.5 * f(abscissa[i-1], ordinates[i-1]) - 0.5 * f(abscissa[i-2], ordinates[i-2]))
    ordinates.append(y_curr)

delta = [abs(ordinates[i] - solution[i]) for i in range(len(abscissa))]

t_prev = t0
y_prev = u0
h = 2 * h
k1 = h * f(t_prev, y_prev)
k2 = h * f(t_prev + h, y_prev + h * k1)
y_curr = y_prev + 0.5 * (k1 + k2)
t_curr = t_prev + h

abscissa2 = [a + i * h for i in range(len(abscissa))]
solution_2_new = [u(x) for x in abscissa2]

ordinates2 = [y_prev, y_curr]

for i in range(2, len(abscissa2)):
    y_curr = y_curr + h * (1.5 * f(abscissa2[i-1], ordinates2[i-1]) - 0.5 * f(abscissa2[i-2], ordinates2[i-2]))
    ordinates2.append(y_curr)

delta2 = [abs(ordinates2[i] - solution_2_new[i]) for i in range(len(abscissa2)) if i % 2 == 0]

workbook = Workbook()
sheet1 = workbook.active
sheet1.title = "Пример 1"

sheet1.cell(row=1, column=1).value = "X"
sheet1.cell(row=1, column=2).value = "Y"
sheet1.cell(row=1, column=3).value = "Точное"
sheet1.cell(row=1, column=4).value = "delta"
sheet1.cell(row=1, column=6).value = "X"
sheet1.cell(row=1, column=7).value = "Y"
sheet1.cell(row=1, column=8).value = "Точное"
sheet1.cell(row=1, column=9).value = "delta"

for i in range(1, len(abscissa) + 1):
    sheet1.cell(row=i + 1, column=1).value = abscissa[i - 1]
    sheet1.cell(row=i + 1, column=2).value = ordinates[i - 1]
    sheet1.cell(row=i + 1, column=3).value = solution[i - 1]
    sheet1.cell(row=i + 1, column=4).value = abs(ordinates[i - 1] - solution[i - 1])

    if i <= len(abscissa) // 2 + 1:
        sheet1.cell(row=i + 1, column=6).value = abscissa2[i - 1]
        sheet1.cell(row=i + 1, column=7).value = ordinates2[i - 1]
        sheet1.cell(row=i + 1, column=8).value = solution_2_new[i - 1]
        sheet1.cell(row=i + 1, column=9).value = abs(ordinates2[i - 1] - solution_2_new[i - 1])

    if i == 54:
        sheet1.cell(row=i + 1, column=16).value = f"max delta (при h = {h / 2})"
        sheet1.cell(row=i + 1, column=19).value = f"max delta (при h = {h})"
        sheet1.cell(row=i + 1, column=22).value = "Отношение погрешностей"
    if i == 55:
        sheet1.cell(row=i + 1, column=17).value = max(delta)
        sheet1.cell(row=i + 1, column=20).value = max(delta2)
        sheet1.cell(row=i + 1, column=23).value = max(delta2) / max(delta)
    if i == 57:
            sheet1.cell(row=i + 1, column=16).value = f"средн. delta (при h = {h / 2})"
            sheet1.cell(row=i + 1, column=19).value = f"средн. delta (при h = {h})"
            sheet1.cell(row=i + 1, column=22).value = "Отношение погрешностей"
    if i == 58:
            sheet1.cell(row=i + 1, column=17).value = average(delta)
            sheet1.cell(row=i + 1, column=20).value = average(delta2)
            sheet1.cell(row=i + 1, column=23).value = average(delta2) / average(delta)

workbook.save("Adams8.xlsx")